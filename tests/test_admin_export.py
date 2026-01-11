from __future__ import annotations

import io

from django.test import TestCase
from django.urls import reverse
from django.contrib.admin.utils import label_for_field

from openpyxl import load_workbook

from django.contrib import admin

from core.admin import CategoryAdmin, EventAdmin
from core.models import Category, Event
from .utils import create_category, create_event, create_user


class AdminExportXlsxTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.super_password = "adminpass12345"
        cls.admin = create_user(username="admin", password=cls.super_password, is_staff=True, is_superuser=True)

        # данные для экспорта
        cls.cat = create_category("ЭкспортКат")
        cls.event = create_event(category=cls.cat, title="ЭкспортСобытие")

    def test_export_requires_admin_login(self):
        url = reverse("admin:export_xlsx")
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 302)
        self.assertIn("/admin/login/", resp.url)

    def test_export_xlsx_generates_valid_workbook(self):
        self.client.login(username=self.admin.username, password=self.super_password)

        url = reverse("admin:export_xlsx")
        resp = self.client.post(url, data={"models": ["core.Category", "core.Event"]})
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("attachment; filename=", resp["Content-Disposition"])

        wb = load_workbook(io.BytesIO(resp.content))
        self.assertIn("Category", wb.sheetnames)
        self.assertIn("Event", wb.sheetnames)

        # Проверяем заголовки по list_display (как в админке)
        ws_cat = wb["Category"]
        cat_admin = CategoryAdmin(Category, admin.site)  # ModelAdmin не требует site для label_for_field
        expected_cat_headers = [label_for_field(f, Category, cat_admin) for f in ("id", "name", "created_at", "updated_at")]
        actual_cat_headers = [cell.value for cell in next(ws_cat.iter_rows(min_row=1, max_row=1))]
        self.assertEqual(actual_cat_headers, expected_cat_headers)

        ws_event = wb["Event"]
        event_admin = EventAdmin(Event, admin.site)
        expected_event_headers = [label_for_field(f, Event, event_admin) for f in ("id", "title", "category", "event_date", "location", "created_at")]
        actual_event_headers = [cell.value for cell in next(ws_event.iter_rows(min_row=1, max_row=1))]
        self.assertEqual(actual_event_headers, expected_event_headers)

        # Проверяем, что выгрузились строки с данными (минимум 1 строка + header)
        self.assertGreaterEqual(ws_cat.max_row, 2)
        self.assertGreaterEqual(ws_event.max_row, 2)
