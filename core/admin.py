from __future__ import annotations

from typing import Any

from django.contrib import admin
from django.contrib.admin.utils import label_for_field, lookup_field
from django.core.exceptions import ObjectDoesNotExist
from django.db import models
from django.http import HttpRequest, HttpResponse
from django.template.response import TemplateResponse
from django.urls import path
from django.utils import formats

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .forms import AdminExportForm
from .models import Category, Event, VolunteerApplication, EventLike


@admin.register(Category)
class CategoryAdmin(admin.ModelAdmin):
    list_display = ("id", "name", "created_at", "updated_at")
    search_fields = ("name",)


@admin.register(Event)
class EventAdmin(admin.ModelAdmin):
    list_display = ("id", "title", "category", "event_date", "location", "created_at")
    list_filter = ("category",)
    search_fields = ("title", "location")


@admin.register(VolunteerApplication)
class VolunteerApplicationAdmin(admin.ModelAdmin):
    list_display = ("id", "user", "event", "status", "created_at")
    list_filter = ("status", "event")
    search_fields = ("user__username", "event__title")


@admin.register(EventLike)
class EventLikeAdmin(admin.ModelAdmin):
    list_display = ("id", "user", "event", "created_at")
    search_fields = ("user__username", "event__title")


def _format_admin_value(value: Any) -> Any:
    if value is None:
        return ""

    if isinstance(value, bool):
        return "Да" if value else "Нет"

    if isinstance(value, models.Model):
        return str(value)

    # datetime/date: локализованный формат как в админке
    if hasattr(value, "tzinfo") and hasattr(value, "year") and hasattr(value, "month"):
        try:
            return formats.date_format(value, format="DATETIME_FORMAT", use_l10n=True)
        except Exception:
            return value.isoformat(sep=" ", timespec="seconds")

    if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day") and not hasattr(value, "hour"):
        try:
            return formats.date_format(value, format="DATE_FORMAT", use_l10n=True)
        except Exception:
            return value.isoformat()

    return value


def _get_admin_columns_and_headers(
    request: HttpRequest,
    model_admin: admin.ModelAdmin,
) -> tuple[list[str], list[str]]:
    columns = list(model_admin.get_list_display(request))
    headers: list[str] = []

    for col in columns:
        try:
            header = label_for_field(col, model_admin.model, model_admin=model_admin, return_attr=False)
        except Exception:
            header = col
        headers.append(str(header))

    return columns, headers


def _get_admin_row_values(
    model_admin: admin.ModelAdmin,
    obj: models.Model,
    columns: list[str],
) -> list[Any]:
    row: list[Any] = []

    for col in columns:
        try:
            _, _, value = lookup_field(col, obj, model_admin)
        except ObjectDoesNotExist:
            value = ""
        except Exception:
            value = ""

        row.append(_format_admin_value(value))

    return row


def export_xlsx_view(request: HttpRequest) -> HttpResponse:
    """
    Экспорт XLSX по колонкам list_display (как в админке).
    Доступ: staff/superuser.
    """
    if not request.user.is_authenticated or not request.user.is_staff:
        return HttpResponse("Forbidden", status=403)

    # Берём текущие ModelAdmin из стандартного admin.site
    model_admin_map: dict[str, admin.ModelAdmin] = {
        "core.Category": admin.site._registry.get(Category),
        "core.Event": admin.site._registry.get(Event),
        "core.VolunteerApplication": admin.site._registry.get(VolunteerApplication),
        "core.EventLike": admin.site._registry.get(EventLike),
    }
    model_admin_map = {k: v for k, v in model_admin_map.items() if v is not None}

    form = AdminExportForm()
    form.fields["models"].choices = [(k, k) for k in model_admin_map.keys()]

    if request.method == "POST":
        form = AdminExportForm(request.POST)
        form.fields["models"].choices = [(k, k) for k in model_admin_map.keys()]

        if form.is_valid():
            selected_models = form.cleaned_data["models"]

            wb = Workbook()
            default_ws = wb.active
            wb.remove(default_ws)

            for model_label in selected_models:
                model_admin = model_admin_map.get(model_label)
                if not model_admin:
                    continue

                # (опционально) строгая проверка прав на модель
                app_label = model_admin.model._meta.app_label
                model_name = model_admin.model._meta.model_name
                perm = f"{app_label}.view_{model_name}"
                if not (request.user.is_superuser or request.user.has_perm(perm)):
                    continue

                columns, headers = _get_admin_columns_and_headers(request, model_admin)

                ws = wb.create_sheet(title=model_label.split(".")[-1][:31])
                ws.append(headers)

                qs = model_admin.get_queryset(request).order_by("id")[:5000]
                for obj in qs:
                    ws.append(_get_admin_row_values(model_admin, obj, columns))

                for i, header in enumerate(headers, start=1):
                    ws.column_dimensions[get_column_letter(i)].width = max(12, min(45, len(str(header)) + 6))

            resp = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            resp["Content-Disposition"] = 'attachment; filename="export.xlsx"'
            wb.save(resp)
            return resp

    return TemplateResponse(request, "admin/export_xlsx.html", {"form": form})


# ✅ Главное: НЕ подменяем admin.site целиком.
# Просто добавляем URL в существующий admin.site через обёртку.
_original_get_urls = admin.site.get_urls


def _patched_get_urls():
    urls = _original_get_urls()
    custom = [
        path("export-xlsx/", admin.site.admin_view(export_xlsx_view), name="export_xlsx"),
    ]
    return custom + urls


admin.site.get_urls = _patched_get_urls  # type: ignore[method-assign]
